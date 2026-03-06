"""
=============================================================
ИМПОРТ ТОВАРОВ ИЗ EXCEL В SUPABASE
=============================================================

Как работает:
    
    Excel-таблица (импорт.xlsx)          Supabase (products)
    ┌──────────────────────────┐         ┌─────────────────────┐
    │ B: Название              │───────► │ name                │
    │ D: Базовая цена (x1.7)   │───────► │ base_price          │
    │ E: Цена при 3 чел (x1.5) │──┐      │                     │
    │ F: Цена при 10 чел (x1.3)│──┼────► │ price_tiers (JSON)  │
    │ G: Ссылка на картинку    │───────► │ image_url           │
    └──────────────────────────┘         └─────────────────────┘

    price_tiers формируется так:
    [
        {"min_quantity": 3,  "price": 3459.00},   ← из столбца E
        {"min_quantity": 10, "price": 2997.80}    ← из столбца F
    ]

Запуск:
    pip install openpyxl supabase
    python import_products.py

=============================================================
"""

import json
import math
import openpyxl
from supabase import create_client

# =============================================================
# НАСТРОЙКИ — ЗАПОЛНИ СВОИ ДАННЫЕ
# =============================================================

SUPABASE_URL = "https://cnijfqedtmmvnnogvteo.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNuaWpmcWVkdG1tdm5ub2d2dGVvIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MDc1OTgzOSwiZXhwIjoyMDg2MzM1ODM5fQ.NfzHhsFSBpRyfQL7dj-wl6soL67AzwcfHfVH6siuoWU"

# Путь к Excel-файлу (поменяй если нужно)
EXCEL_FILE = "импорт.xlsx"

# ID категории "Косметика" в Supabase (проверь в своей БД)
CATEGORY_ID = 2

# Начальный остаток на складе для всех товаров
DEFAULT_STOCK = 100


# =============================================================
# ОСНОВНАЯ ЛОГИКА
# =============================================================

def round_price(value):
    """Округляем цену до целых рублей вверх.
    
    Пример: 791.35 → 792
    """
    return math.ceil(value)


def read_excel(filepath):
    """Читаем Excel и достаём данные о товарах.
    
    Представь это как конвейер на фабрике:
    каждая строка Excel — это коробка с товаром,
    мы открываем каждую и аккуратно раскладываем
    содержимое по полочкам (словарь).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Читаем min_quantity из заголовков E1 и F1
    tier1_qty = int(ws["E1"].value)  # 3
    tier2_qty = int(ws["F1"].value)  # 10
    
    products = []
    
    # Идём по строкам начиная со 2-й (1-я — заголовки)
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        barcode, name, purchase_price, base_price, tier1_price, tier2_price, image_url = row
        
        # Пропускаем пустые строки
        if not name or not purchase_price:
            continue
        
        # Если формулы не вычислились (data_only не сработал),
        # считаем цены вручную из закупочной
        if base_price is None or isinstance(base_price, str):
            base_price = purchase_price * 1.7
        if tier1_price is None or isinstance(tier1_price, str):
            tier1_price = purchase_price * 1.5
        if tier2_price is None or isinstance(tier2_price, str):
            tier2_price = purchase_price * 1.3
        
        product = {
            "name": name.strip(),
            "base_price": round_price(base_price),
            "price_tiers": [
                {"min_quantity": tier1_qty, "price": round_price(tier1_price)},
                {"min_quantity": tier2_qty, "price": round_price(tier2_price)},
            ],
            "image_url": image_url.strip() if image_url else None,
            "category_id": CATEGORY_ID,
            "stock": DEFAULT_STOCK,
            "is_active": True,
        }
        
        products.append(product)
    
    return products


def upload_to_supabase(products):
    """Загружаем товары в Supabase.
    
    Представь, что мы пришли в магазин с тележкой
    товаров и расставляем их на полки (в базу данных).
    Каждый товар кладётся отдельно — если один не влез,
    остальные всё равно встанут на место.
    """
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    success = 0
    errors = []
    
    for i, product in enumerate(products, 1):
        try:
            # Supabase Python SDK автоматически сериализует dict → JSON
            result = client.table("products").insert(product).execute()
            
            success += 1
            print(f"  ✅ [{i}/{len(products)}] {product['name'][:50]}...")
            
        except Exception as e:
            errors.append({"name": product["name"], "error": str(e)})
            print(f"  ❌ [{i}/{len(products)}] {product['name'][:50]}... — ОШИБКА: {e}")
    
    return success, errors


def main():
    print("=" * 60)
    print("  ИМПОРТ ТОВАРОВ В SUPABASE")
    print("=" * 60)
    
    # Шаг 1: Читаем Excel
    print(f"\n📄 Читаю файл: {EXCEL_FILE}")
    products = read_excel(EXCEL_FILE)
    print(f"   Найдено товаров: {len(products)}")
    
    # Шаг 2: Показываем превью первых 3 товаров
    print("\n📋 Превью (первые 3 товара):")
    for p in products[:3]:
        print(f"   • {p['name'][:60]}")
        print(f"     Базовая: {p['base_price']}₽ | "
              f"От {p['price_tiers'][0]['min_quantity']} чел: {p['price_tiers'][0]['price']}₽ | "
              f"От {p['price_tiers'][1]['min_quantity']} чел: {p['price_tiers'][1]['price']}₽")
    
    # Шаг 3: Спрашиваем подтверждение
    print(f"\n⚠️  Будет загружено {len(products)} товаров в Supabase")
    confirm = input("   Продолжить? (y/n): ").strip().lower()
    
    if confirm != "y":
        print("\n❌ Импорт отменён")
        return
    
    # Шаг 4: Загружаем
    print("\n🚀 Загружаю в Supabase...")
    success, errors = upload_to_supabase(products)
    
    # Шаг 5: Итоги
    print("\n" + "=" * 60)
    print(f"  ИТОГО: ✅ {success} загружено | ❌ {len(errors)} ошибок")
    print("=" * 60)
    
    if errors:
        print("\n⚠️  Товары с ошибками:")
        for err in errors:
            print(f"   • {err['name'][:50]} — {err['error']}")


if __name__ == "__main__":
    main()
